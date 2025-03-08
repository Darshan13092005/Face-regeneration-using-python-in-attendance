import face_recognition
import cv2
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

students_folder = r"D:\student"
image_extensions = [".jpg", ".jpeg", ".png"]
known_face_encodings = []
known_face_names = []

if not os.path.exists(students_folder):
    print("Error: The folder for student images does not exist.")
    exit()

for filename in os.listdir(students_folder):
    if any(filename.lower().endswith(ext) for ext in image_extensions):
        image_path = os.path.join(students_folder, filename)
        student_name = os.path.splitext(filename)[0]
        image = face_recognition.load_image_file(image_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            known_face_encodings.append(encodings[0])
            known_face_names.append(student_name)

attendance = {}
attendance_file = "attendance.xlsx"
if not os.path.exists(attendance_file):
    wb = Workbook()
    wb.save(attendance_file)

video_capture = cv2.VideoCapture(0)
if not video_capture.isOpened():
    print("Error: Unable to access the webcam.")
    exit()

while True:
    ret, frame = video_capture.read()
    if not ret:
        print("Error: Unable to capture video. Exiting...")
        break
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
    for face_encoding, face_location in zip(face_encodings, face_locations):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"
        if True in matches:
            match_index = matches.index(True)
            name = known_face_names[match_index]
            now = datetime.now()
            attendance[name] = {
                "Date": now.strftime("%Y-%m-%d"),
                "Time": now.strftime("%H:%M:%S"),
            }
            print(f"Attendance updated for: {name} at {attendance[name]['Time']}")
        top, right, bottom, left = [v * 4 for v in face_location]
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    cv2.imshow("Face Recognition Attendance System", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

if attendance:
    today_date = datetime.now().strftime("%Y-%m-%d")
    try:
        wb = load_workbook(attendance_file)
        if today_date not in wb.sheetnames:
            ws = wb.create_sheet(title=today_date)
            ws.append(["Name", "Date", "Time"])
        else:
            ws = wb[today_date]
        existing_records = {ws.cell(row, 1).value: row for row in range(2, ws.max_row + 1)}
        for name, details in attendance.items():
            if name in existing_records:
                row = existing_records[name]
                ws.cell(row=row, column=2, value=details["Date"])
                ws.cell(row=row, column=3, value=details["Time"])
            else:
                ws.append([name, details["Date"], details["Time"]])
        wb.save(attendance_file)
        print(f"Attendance for {today_date} saved in sheet: {today_date}")
    except Exception as e:
        print(f"Error saving attendance: {e}")
else:
    print("No attendance recorded.")

video_capture.release()
cv2.destroyAllWindows()
